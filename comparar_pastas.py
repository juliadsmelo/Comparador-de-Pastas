
import os
from collections import defaultdict

# Necessário instalar: pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def listar_arquivos_relativos(pasta_raiz):
    """Lista todos os arquivos e subpastas com caminhos relativos."""
    arquivos = set()
    for dirpath, dirnames, filenames in os.walk(pasta_raiz):
        rel_dir = os.path.relpath(dirpath, pasta_raiz)

        for dirname in dirnames:
            caminho_relativo = os.path.join(rel_dir, dirname)
            arquivos.add(("PASTA", caminho_relativo))

        for filename in filenames:
            caminho_relativo = os.path.join(rel_dir, filename)
            arquivos.add(("ARQUIVO", caminho_relativo))

    return arquivos


def pasta_esta_vazia(pasta_raiz, caminho_relativo_pasta):
    """Verifica se uma pasta está vazia (sem arquivos, inclusive em subpastas)."""
    caminho_completo = os.path.join(pasta_raiz, caminho_relativo_pasta)
    for dirpath, dirnames, filenames in os.walk(caminho_completo):
        if filenames:
            return False
    return True


def agrupar_por_pasta(itens):
    """Agrupa itens faltantes pela pasta-pai."""
    agrupado = defaultdict(list)
    for tipo, caminho in itens:
        pasta_pai = os.path.dirname(caminho)
        nome = os.path.basename(caminho)
        if pasta_pai == ".":
            pasta_pai = "(raiz)"
        agrupado[pasta_pai].append((tipo, nome))
    return agrupado


def exportar_excel(itens_faltando, caminho_excel):
    """Exporta os itens faltantes para um arquivo Excel formatado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Itens Faltantes"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cabeçalhos
    headers = ["Pasta", "Tipo", "Nome do Item"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda

    # Dados
    agrupado = agrupar_por_pasta(itens_faltando)
    row = 2
    for pasta in sorted(agrupado.keys()):
        itens = sorted(agrupado[pasta], key=lambda x: (x[0], x[1]))
        for tipo, nome in itens:
            ws.cell(row=row, column=1, value=pasta).border = borda
            ws.cell(row=row, column=2, value=tipo).border = borda
            ws.cell(row=row, column=3, value=nome).border = borda
            row += 1

    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50

    # Filtros automáticos
    ws.auto_filter.ref = f"A1:C{row - 1}"

    wb.save(caminho_excel)


def comparar_pastas(pasta_antiga, pasta_nova, caminho_excel):
    """Compara duas pastas, exibe resumo agrupado e exporta Excel."""

    print("=" * 60)
    print("  RELATÓRIO DE COMPARAÇÃO DE PASTAS")
    print("=" * 60)
    print(f"\n  Referência (antiga): {pasta_antiga}")
    print(f"  Verificação (nova):  {pasta_nova}\n")

    # Listar conteúdo
    conteudo_antiga = listar_arquivos_relativos(pasta_antiga)
    conteudo_nova = listar_arquivos_relativos(pasta_nova)

    # Itens faltando (antes da filtragem)
    faltando_bruto = conteudo_antiga - conteudo_nova

    # Filtrar: remover pastas que estão vazias na pasta antiga
    faltando = set()
    for tipo, caminho in faltando_bruto:
        if tipo == "PASTA":
            # Só inclui a pasta se ela NÃO estiver vazia na origem
            if not pasta_esta_vazia(pasta_antiga, caminho):
                faltando.add((tipo, caminho))
        else:
            faltando.add((tipo, caminho))

    if not faltando:
        print("  ✅ Nenhum item faltando! A pasta nova está completa.")
        print("     (Pastas vazias foram desconsideradas)")
        return

    # Agrupar por pasta
    agrupado = agrupar_por_pasta(faltando)

    # Exibir resumo agrupado no terminal
    print("-" * 60)
    print("  RESUMO POR PASTA (itens faltantes):")
    print("-" * 60)

    total_pastas = 0
    total_arquivos = 0

    for pasta in sorted(agrupado.keys()):
        itens = agrupado[pasta]
        qtd_pastas = sum(1 for t, _ in itens if t == "PASTA")
        qtd_arquivos = sum(1 for t, _ in itens if t == "ARQUIVO")
        total_pastas += qtd_pastas
        total_arquivos += qtd_arquivos

        print(f"\n  📂 {pasta}")
        if qtd_pastas > 0:
            print(f"     └─ {qtd_pastas} subpasta(s) faltando")
        if qtd_arquivos > 0:
            print(f"     └─ {qtd_arquivos} arquivo(s) faltando")

    # Resumo final
    print("\n" + "=" * 60)
    print("  TOTAIS:")
    print(f"     📁 Subpastas faltando:  {total_pastas}")
    print(f"     📄 Arquivos faltando:   {total_arquivos}")
    print(f"     🔢 Total de itens:      {len(faltando)}")
    print("=" * 60)
    print("\n  ℹ️  Pastas vazias foram desconsideradas da análise.")

    # Exportar Excel
    exportar_excel(faltando, caminho_excel)
    print(f"\n  📊 Detalhes exportados para: {caminho_excel}")
    print("     (Abra o Excel para ver todos os itens com filtros)\n")


# ============================================================
# SUBSTITUA OS CAMINHOS ABAIXO PELOS SEUS:
# ============================================================

pasta_antiga = r"C:\Users\GFE0601\OneDrive - MDLZ\Engineering Curitiba - Documentos\06. Infra & Digital Projects"  # ← SUBSTITUA AQUI
pasta_nova = r"C:\Users\GFE0601\OneDrive - MDLZ\Engineering Curitiba - Documentos\06. Infra&Digital Projects"      # ← SUBSTITUA AQUI
caminho_excel = r"C:\Users\GFE0601\OneDrive - MDLZ\Área de Trabalho\Relatório faltantes\itens faltantes.xlsx"            # ← ONDE SALVAR O EXCEL

comparar_pastas(pasta_antiga, pasta_nova, caminho_excel)

